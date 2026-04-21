from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from dotenv import dotenv_values
from openpyxl import load_workbook
from sqlalchemy import URL, bindparam, create_engine, text

if TYPE_CHECKING:
    from collections.abc import Iterable

    from sqlalchemy.engine import Engine


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / '合并结果.xlsx'
DEFAULT_ENV_FILE = BASE_DIR / '.env.dev'
DEFAULT_SHEET_NAME = 'Sheet1'
TARGET_TABLE = 'std_standard_content'

COLUMN_MAP = {
    '标准号': 'standard_no',
    '同级内排序号': 'sibling_order_no',
    '内容层级': 'content_level',
    '所在文件页面': 'source_page_no',
    '内容类型': 'content_type',
    '内容': 'content_text_with_no',
}
REQUIRED_HEADERS = list(COLUMN_MAP)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='临时导入合并结果.xlsx 的标准正文内容到 MySQL')
    parser.add_argument('--workbook', type=Path, default=DEFAULT_WORKBOOK, help='待导入的 Excel 文件路径')
    parser.add_argument('--sheet', default=DEFAULT_SHEET_NAME, help='待导入的工作表名称')
    parser.add_argument('--env-file', type=Path, default=DEFAULT_ENV_FILE, help='.env.dev 配置文件路径')
    parser.add_argument('--batch-size', type=int, default=2000, help='批量插入条数')
    parser.add_argument('--commit', action='store_true', help='实际写入数据库；不加该参数时只预览')
    parser.add_argument(
        '--replace',
        action='store_true',
        help='写入前删除 Excel 中同一批标准号在目标表里的已有内容；必须配合 --commit 使用',
    )
    parser.add_argument('--check-db', action='store_true', help='预览时也连接数据库检查目标表与当前最大 id')
    parser.add_argument('--created-by', default='temp_import_standard_content', help='created_by/updated_by 字段值')
    parser.add_argument('--start-id', type=int, default=None, help='手动指定导入起始 id；默认使用目标表 max(id)+1')
    return parser.parse_args()


def blank_to_none(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def to_text(value: Any) -> str | None:
    value = blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_int(value: Any, row_no: int, column_name: str) -> int | None:
    value = blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f'第 {row_no} 行“{column_name}”不是有效整数: {value!r}')
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text_value = str(value).strip()
    try:
        number = float(text_value)
    except ValueError as exc:
        raise ValueError(f'第 {row_no} 行“{column_name}”不是有效整数: {value!r}') from exc
    if not number.is_integer():
        raise ValueError(f'第 {row_no} 行“{column_name}”不是有效整数: {value!r}')
    return int(number)


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ''


def load_rows(workbook_path: Path, sheet_name: str) -> tuple[list[dict[str, Any]], int]:
    workbook_path = workbook_path.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f'找不到 Excel 文件: {workbook_path}')

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'找不到工作表 {sheet_name!r}，当前工作表: {workbook.sheetnames}')

        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError(f'工作表 {sheet_name!r} 没有表头')

        headers = [normalize_header(value) for value in header_row]
        missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            raise ValueError(f'工作表缺少列: {", ".join(missing_headers)}')

        column_indexes = {header: headers.index(header) for header in REQUIRED_HEADERS}
        rows: list[dict[str, Any]] = []
        skipped_empty = 0
        invalid_messages: list[str] = []

        for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_values = {header: row[column_indexes[header]] for header in REQUIRED_HEADERS}
            if all(blank_to_none(value) is None for value in raw_values.values()):
                skipped_empty += 1
                continue

            try:
                record = {
                    'id': None,
                    'standard_no': to_text(raw_values['标准号']),
                    'parent_node_id': None,
                    'sibling_order_no': to_int(raw_values['同级内排序号'], row_no, '同级内排序号'),
                    'content_level': to_int(raw_values['内容层级'], row_no, '内容层级'),
                    'source_page_no': to_int(raw_values['所在文件页面'], row_no, '所在文件页面'),
                    'clause_no': None,
                    'content_type': to_text(raw_values['内容类型']),
                    'content_text_with_no': to_text(raw_values['内容']),
                    'is_mandatory': None,
                    'created_at': None,
                    'created_by': None,
                    'updated_at': None,
                    'updated_by': None,
                    'is_deleted': 0,
                    'remark': None,
                }
            except ValueError as exc:
                invalid_messages.append(str(exc))
                continue

            if record['standard_no'] is None:
                invalid_messages.append(f'第 {row_no} 行缺少“标准号”')
                continue

            rows.append(record)

        if invalid_messages:
            preview = '\n'.join(invalid_messages[:20])
            omitted = len(invalid_messages) - 20
            suffix = f'\n还有 {omitted} 条错误未显示' if omitted > 0 else ''
            raise ValueError(f'Excel 数据存在无效行:\n{preview}{suffix}')

        return rows, skipped_empty
    finally:
        workbook.close()


def load_mysql_engine(env_file: Path) -> Engine:
    env_file = env_file.resolve()
    if not env_file.exists():
        raise FileNotFoundError(f'找不到环境配置文件: {env_file}')

    env = dotenv_values(env_file)
    db_type = str(env.get('DB_TYPE', 'mysql')).strip().lower()
    if db_type != 'mysql':
        raise ValueError(f'当前脚本只支持 MySQL，.env 配置的 DB_TYPE 是 {db_type!r}')

    missing = [key for key in ['DB_HOST', 'DB_PORT', 'DB_USERNAME', 'DB_PASSWORD', 'DB_DATABASE'] if not env.get(key)]
    if missing:
        raise ValueError(f'.env 缺少数据库配置: {", ".join(missing)}')

    url = URL.create(
        'mysql+pymysql',
        username=str(env['DB_USERNAME']),
        password=str(env['DB_PASSWORD']),
        host=str(env['DB_HOST']),
        port=int(str(env['DB_PORT'])),
        database=str(env['DB_DATABASE']),
        query={'charset': 'utf8mb4'},
    )
    return create_engine(url, future=True, pool_pre_ping=True)


def batches(items: list[dict[str, Any]], size: int) -> Iterable[list[dict[str, Any]]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]


def assign_runtime_fields(
    rows: list[dict[str, Any]],
    start_id: int,
    now: datetime,
    operator: str,
) -> None:
    for offset, row in enumerate(rows):
        row['id'] = start_id + offset
        row['created_at'] = now
        row['created_by'] = operator
        row['updated_at'] = now
        row['updated_by'] = operator


def fetch_next_id(engine: Engine) -> int:
    with engine.connect() as connection:
        max_id = connection.execute(text(f'SELECT COALESCE(MAX(id), 0) FROM {TARGET_TABLE}')).scalar_one()
    return int(max_id) + 1


def replace_existing_rows(engine: Engine, standard_nos: list[str]) -> int:
    delete_sql = text(f'DELETE FROM {TARGET_TABLE} WHERE standard_no IN :standard_nos').bindparams(
        bindparam('standard_nos', expanding=True),
    )
    deleted_count = 0
    with engine.begin() as connection:
        for standard_no_batch in batches([{'standard_no': value} for value in standard_nos], 500):
            values = [item['standard_no'] for item in standard_no_batch]
            result = connection.execute(delete_sql, {'standard_nos': values})
            deleted_count += result.rowcount or 0
    return deleted_count


def insert_rows(engine: Engine, rows: list[dict[str, Any]], batch_size: int) -> int:
    insert_sql = text(
        f"""
        INSERT INTO {TARGET_TABLE} (
            id,
            standard_no,
            parent_node_id,
            sibling_order_no,
            content_level,
            source_page_no,
            clause_no,
            content_type,
            content_text_with_no,
            is_mandatory,
            created_at,
            created_by,
            updated_at,
            updated_by,
            is_deleted,
            remark
        ) VALUES (
            :id,
            :standard_no,
            :parent_node_id,
            :sibling_order_no,
            :content_level,
            :source_page_no,
            :clause_no,
            :content_type,
            :content_text_with_no,
            :is_mandatory,
            :created_at,
            :created_by,
            :updated_at,
            :updated_by,
            :is_deleted,
            :remark
        )
        """
    )
    inserted_count = 0
    with engine.begin() as connection:
        for row_batch in batches(rows, batch_size):
            result = connection.execute(insert_sql, row_batch)
            inserted_count += result.rowcount or len(row_batch)
    return inserted_count


def main() -> None:
    args = parse_args()
    if args.batch_size <= 0:
        raise ValueError('--batch-size 必须大于 0')
    if args.replace and not args.commit:
        raise ValueError('--replace 必须配合 --commit 使用')

    rows, skipped_empty = load_rows(args.workbook, args.sheet)
    standard_nos = sorted({row['standard_no'] for row in rows if row['standard_no']})
    print(f'Excel 文件: {args.workbook.resolve()}')
    print(f'工作表: {args.sheet}')
    print(f'有效待导入行数: {len(rows)}')
    print(f'涉及标准号数量: {len(standard_nos)}')
    print(f'跳过空行数: {skipped_empty}')

    if not args.commit and not args.check_db:
        print('当前为预览模式，未连接数据库、未写入。加 --commit 后执行入库。')
        return

    engine = load_mysql_engine(args.env_file)
    next_id = args.start_id if args.start_id is not None else fetch_next_id(engine)
    print(f'目标表: {TARGET_TABLE}')
    print(f'导入起始 id: {next_id}')

    if not args.commit:
        print('当前为预览模式，已完成数据库连通性与 max(id) 检查，未写入。')
        return

    now = datetime.now()
    assign_runtime_fields(rows, next_id, now, args.created_by)

    deleted_count = 0
    if args.replace:
        deleted_count = replace_existing_rows(engine, standard_nos)
        print(f'已删除目标表中同批标准号旧数据: {deleted_count}')

    inserted_count = insert_rows(engine, rows, args.batch_size)
    print(f'导入完成，插入行数: {inserted_count}')


if __name__ == '__main__':
    main()
